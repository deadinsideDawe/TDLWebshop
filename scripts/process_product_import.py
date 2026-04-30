from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = Path(r"C:\Users\Dell\Downloads\termek-import-azonnali-feltolteshez.xlsx")
OUTPUT_DIR = REPO_ROOT / "docs" / "generated-import"
OUTPUT_XLSX = OUTPUT_DIR / "termek-import-vegleges.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "termek-import-vegleges.csv"
REPORT_JSON = OUTPUT_DIR / "termek-import-report.json"

EXPECTED_HEADERS = [
    "name",
    "price",
    "category",
    "image",
    "galleryImages",
    "stock",
    "stockQuantity",
    "sku",
    "brand",
    "shortDescription",
    "description",
    "isWeeklyDeal",
    "isTopProduct",
    "salePercent",
    "saleStartsAt",
    "saleEndsAt",
]

CATEGORY_CODES = {
    "Fűtés": "FUT",
    "Hűtés": "HUT",
    "Víz": "VIZ",
    "Szellőzés": "SZEL",
    "Szerelvények": "SZER",
    "Lakossági megoldások": "LAK",
}

EDITABLE_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
PATH_FIXES = {
    "product/futes/hajdu_pt_1.jpg": "products/futes/hajdu_pt_1.jpg",
    "products/gutes/hajdu_pt_2.jpg": "products/futes/hajdu_pt_2.jpg",
    "products/futes/install_calidum_fo.jpg": "products/futes/intall_calidum_fo.jpg",
    "produts/futes/emmeti_fo.jpg": "products/futes/emmeti_fo.jpg",
    "products/hutes/gree_moma_fo": "products/hutes/gree_moma_fo.jpg",
    "products/szellozes/lg_lz_fo.jpg": "products/szellozes/lg_tz_fo.jpg",
    "products/viz/kolo_rekors_2.jpg": "products/viz/kolo_rekord_2.jpg",
    "products/víz/wellis_dublo_2.jpg": "products/viz/wellis_dublo_2.jpg",
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    without_accents = "".join(char for char in normalized if not unicodedata.combining(char))
    upper_value = without_accents.upper()
    return re.sub(r"[^A-Z0-9]+", "-", upper_value).strip("-")


def normalize_bool(value) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value or "").strip().lower()
    return "true" if text in {"1", "true", "igen", "yes"} else "false"


def normalize_gallery(image: str, gallery_raw: str) -> str:
    parts = [item.strip() for item in str(gallery_raw or "").split("|") if item.strip()]
    unique = []
    for item in [image.strip(), *parts]:
        if item and item not in unique:
            unique.append(item)
    if len(unique) <= 1:
        return ""
    return "|".join(unique[1:])


def normalize_image_path(path: str) -> str:
    normalized = str(path or "").strip().replace("\\", "/")
    if not normalized:
        return ""

    normalized = PATH_FIXES.get(normalized, normalized)
    normalized = normalized.replace("products/víz/", "products/viz/")
    normalized = normalized.replace("produts/", "products/")
    normalized = normalized.replace("product/", "products/")
    return normalized


def infer_stock_quantity(price: float, category: str, name: str) -> int:
    lowered = name.lower()
    if any(keyword in lowered for keyword in ["kazán", "bojler", "klíma", "hővisszanyer", "kád", "wc", "zuhanykabin"]):
        if price >= 300000:
            return 1
        if price >= 150000:
            return 2
        if price >= 70000:
            return 3
        return 4

    if any(keyword in lowered for keyword in ["termosztát", "szelep", "szifon", "csaptelep", "ventilátor", "rács"]):
        if price >= 50000:
            return 5
        if price >= 20000:
            return 8
        return 12

    if category == "Szerelvények":
        if price >= 30000:
            return 8
        return 15

    if category == "Lakossági megoldások":
        if price >= 100000:
            return 3
        if price >= 30000:
            return 6
        return 10

    if price >= 250000:
        return 2
    if price >= 100000:
        return 4
    if price >= 50000:
        return 6
    if price >= 15000:
        return 10
    return 14


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    if not input_path.exists():
        raise FileNotFoundError(f"Nem találom a bemeneti Excel fájlt: {input_path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path)
    sheet = workbook["Termekek"]

    headers = []
    for cell in sheet[1]:
        if cell.value is None:
            headers.append("")
        else:
            headers.append(str(cell.value).strip())

    header_index = {header: idx for idx, header in enumerate(headers) if header}
    missing_headers = [header for header in EXPECTED_HEADERS if header not in header_index]
    if missing_headers:
        raise ValueError(f"Hiányzó kötelező oszlopok: {', '.join(missing_headers)}")

    generated_skus = defaultdict(int)
    rows_for_csv = []
    missing_images = []
    processed_rows = 0

    for row_number in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row_number, column=header_index["name"] + 1).value
        if not str(name or "").strip():
            continue

        processed_rows += 1
        category = str(sheet.cell(row=row_number, column=header_index["category"] + 1).value or "").strip()
        image = normalize_image_path(sheet.cell(row=row_number, column=header_index["image"] + 1).value or "")
        gallery_raw = str(sheet.cell(row=row_number, column=header_index["galleryImages"] + 1).value or "").strip()
        sheet.cell(row=row_number, column=header_index["image"] + 1).value = image
        price_value = sheet.cell(row=row_number, column=header_index["price"] + 1).value
        price = float(price_value)

        normalized_gallery = normalize_gallery(
            image,
            "|".join(normalize_image_path(item) for item in str(gallery_raw or "").split("|"))
        )
        sheet.cell(row=row_number, column=header_index["galleryImages"] + 1).value = normalized_gallery

        stock_qty_cell = sheet.cell(row=row_number, column=header_index["stockQuantity"] + 1)
        stock_qty = stock_qty_cell.value
        if stock_qty in (None, ""):
            stock_qty = infer_stock_quantity(price, category, str(name))
            stock_qty_cell.value = stock_qty
        stock_qty = int(float(stock_qty))

        stock_cell = sheet.cell(row=row_number, column=header_index["stock"] + 1)
        stock_value = str(stock_cell.value or "").strip()
        if not stock_value:
            stock_value = "Készleten" if stock_qty > 0 else "Rendelésre"
            stock_cell.value = stock_value

        sku_cell = sheet.cell(row=row_number, column=header_index["sku"] + 1)
        sku_value = str(sku_cell.value or "").strip()
        if not sku_value:
            category_code = CATEGORY_CODES.get(category, slugify(category)[:4] or "TERM")
            generated_skus[category] += 1
            sku_value = f"TDL-{category_code}-{generated_skus[category]:03d}"
            sku_cell.value = sku_value

        sale_starts = sheet.cell(row=row_number, column=header_index["saleStartsAt"] + 1).value or 0
        sale_ends = sheet.cell(row=row_number, column=header_index["saleEndsAt"] + 1).value or 0
        weekly = normalize_bool(sheet.cell(row=row_number, column=header_index["isWeeklyDeal"] + 1).value)
        top = normalize_bool(sheet.cell(row=row_number, column=header_index["isTopProduct"] + 1).value)
        sale_percent = int(float(sheet.cell(row=row_number, column=header_index["salePercent"] + 1).value or 0))

        repo_image = REPO_ROOT / "public" / image
        if image and not repo_image.exists():
            missing_images.append(image)

        existing_gallery_items = []
        if normalized_gallery:
            for gallery_item in normalized_gallery.split("|"):
                gallery_path = REPO_ROOT / "public" / gallery_item
                if gallery_path.exists():
                    existing_gallery_items.append(gallery_item)
                else:
                    missing_images.append(gallery_item)
            normalized_gallery = "|".join(existing_gallery_items)
            sheet.cell(row=row_number, column=header_index["galleryImages"] + 1).value = normalized_gallery

        row_values = {
            "name": str(name).strip(),
            "price": int(price) if float(price).is_integer() else price,
            "category": category,
            "image": image,
            "galleryImages": normalized_gallery,
            "stock": stock_value,
            "stockQuantity": stock_qty,
            "sku": sku_value,
            "brand": str(sheet.cell(row=row_number, column=header_index["brand"] + 1).value or "").strip(),
            "shortDescription": str(sheet.cell(row=row_number, column=header_index["shortDescription"] + 1).value or "").strip(),
            "description": str(sheet.cell(row=row_number, column=header_index["description"] + 1).value or "").strip(),
            "isWeeklyDeal": weekly,
            "isTopProduct": top,
            "salePercent": sale_percent,
            "saleStartsAt": int(float(sale_starts or 0)),
            "saleEndsAt": int(float(sale_ends or 0)),
        }
        rows_for_csv.append(row_values)

    workbook.save(OUTPUT_XLSX)

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(EXPECTED_HEADERS)
        for row in rows_for_csv:
            writer.writerow([row[header] for header in EXPECTED_HEADERS])

    report = {
        "input": str(input_path),
        "output_xlsx": str(OUTPUT_XLSX),
        "output_csv": str(OUTPUT_CSV),
        "processed_rows": processed_rows,
        "missing_images": sorted(set(missing_images)),
    }
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
