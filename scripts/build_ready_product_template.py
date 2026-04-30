from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "termek-import-azonnali-feltolteshez.xlsx"


HEADERS = [
    "name",
    "price",
    "category",
    "image",
    "galleryImages",
    "stock",
    "stockQuantity",
    "sku",
    "brand",
    "shortDescription",
    "description",
    "isWeeklyDeal",
    "isTopProduct",
    "salePercent",
    "saleStartsAt",
    "saleEndsAt",
]


CATEGORIES = [
    ("Fűtés", "TDL-FUT", 10),
    ("Hűtés", "TDL-HUT", 10),
    ("Víz", "TDL-VIZ", 10),
    ("Szellőzés", "TDL-SZELL", 10),
    ("Szerelvények", "TDL-SZER", 10),
    ("Lakossági megoldások", "TDL-LAK", 10),
]


USER_INPUT_COLUMNS = {
    "A",  # name
    "B",  # price
    "D",  # image
    "E",  # galleryImages
    "G",  # stockQuantity
    "H",  # sku
    "I",  # brand
    "J",  # shortDescription
    "K",  # description
}


def build_workbook() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "Termekek"

    category_sheet = wb.create_sheet("Kategoriak")
    guide_sheet = wb.create_sheet("Kitoltesi_utmutato")

    create_category_sheet(category_sheet)
    create_products_sheet(ws)
    create_guide_sheet(guide_sheet)

    return wb


def create_products_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    fixed_fill = PatternFill("solid", fgColor="E8EEF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    category_fill = PatternFill("solid", fgColor="DDEBF7")
    border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    ws.append(HEADERS)

    row_index = 2
    for category_name, sku_prefix, count in CATEGORIES:
        for item_no in range(1, count + 1):
            ws.append(
                [
                    "",  # name
                    "",  # price
                    category_name,
                    "",  # image
                    "",  # galleryImages
                    "Raktáron",
                    "",  # stockQuantity
                    "",  # sku
                    "",  # brand
                    "",  # shortDescription
                    "",  # description
                    False,
                    False,
                    0,
                    0,
                    0,
                ]
            )
            ws[f"Q{row_index}"] = sku_prefix
            ws[f"R{row_index}"] = item_no
            row_index += 1

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    max_row = ws.max_row
    for row in range(2, max_row + 1):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            col_letter = cell.column_letter
            if col_letter in USER_INPUT_COLUMNS:
                cell.fill = input_fill
            elif col_letter == "C":
                cell.fill = category_fill
            else:
                cell.fill = fixed_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{max_row}"

    widths = {
        "A": 28,
        "B": 12,
        "C": 24,
        "D": 28,
        "E": 40,
        "F": 14,
        "G": 14,
        "H": 16,
        "I": 18,
        "J": 30,
        "K": 48,
        "L": 14,
        "M": 14,
        "N": 12,
        "O": 14,
        "P": 12,
        "Q": 12,
        "R": 8,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 42

    stock_validation = DataValidation(
        type="list",
        formula1='"Raktáron,Rendelésre,Nincs készleten"',
        allow_blank=False,
    )
    category_validation = DataValidation(
        type="list",
        formula1="=Kategoriak!$A$2:$A$7",
        allow_blank=False,
    )
    boolean_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
    )

    ws.add_data_validation(stock_validation)
    ws.add_data_validation(category_validation)
    ws.add_data_validation(boolean_validation)

    stock_validation.add(f"F2:F{max_row}")
    category_validation.add(f"C2:C{max_row}")
    boolean_validation.add(f"L2:M{max_row}")

    table = Table(displayName="TermekImportTabla", ref=f"A1:P{max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    ws["T1"] = "Megjegyzés"
    ws["T2"] = "Sárga mezők = te töltöd ki"
    ws["T3"] = "Kék/szürke mezők = előre kitöltve"
    ws["T4"] = "galleryImages mezőben több kép | jellel elválasztva"
    ws["T5"] = "Példa SKU: TDL-FUT-001"
    ws["T6"] = "sale mezők maradhatnak 0-n, ha nincs akció"
    ws.column_dimensions["T"].width = 44


def create_category_sheet(ws):
    ws["A1"] = "Kategória"
    ws["B1"] = "SKU előtag"
    ws["C1"] = "Ajánlott darabszám"

    for idx, (category_name, sku_prefix, count) in enumerate(CATEGORIES, start=2):
        ws[f"A{idx}"] = category_name
        ws[f"B{idx}"] = sku_prefix
        ws[f"C{idx}"] = count

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18


def create_guide_sheet(ws):
    rows = [
        ("Mit kell kitöltened?", "Csak a sárgával jelölt mezőket."),
        ("Kötelező mezők", "name, price, image, stockQuantity, sku, brand, shortDescription, description"),
        ("Több kép megadása", "A galleryImages mezőbe írd a képeket | jellel elválasztva."),
        ("Fő kép példa", "products/radiator-szelep-fo.jpg"),
        ("Galéria példa", "products/radiator-szelep-1.jpg|products/radiator-szelep-2.jpg"),
        ("SKU példa", "TDL-FUT-001 vagy TDL-VIZ-014"),
        ("Stock mező", "Alapból Raktáron. Ha kell, átírhatod Rendelésre vagy Nincs készleten értékre."),
        ("Akció", "Ha nincs akció, hagyd a salePercent / saleStartsAt / saleEndsAt mezőket 0 értéken."),
        ("Kiemelés", "Az isWeeklyDeal és isTopProduct alapból FALSE. Ha kell, állítsd TRUE-ra."),
        ("Kategóriák", ", ".join(category for category, _, _ in CATEGORIES)),
    ]

    ws["A1"] = "Téma"
    ws["B1"] = "Magyarázat"
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


if __name__ == "__main__":
    workbook = build_workbook()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(OUTPUT)
